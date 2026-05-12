import os
import sys
import itertools
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font
import rootutils

from pathlib import Path
POCKET_ROOT = Path(__file__).resolve().parent.parent.parent  # Pocket/
try:
    ROOT = rootutils.setup_root(__file__, indicator=".project-root", pythonpath=True)
except Exception:
    ROOT = POCKET_ROOT
    if str(POCKET_ROOT) not in sys.path:
        sys.path.insert(0, str(POCKET_ROOT))
if str(POCKET_ROOT) in sys.path:
    sys.path.remove(str(POCKET_ROOT))
sys.path.insert(0, str(POCKET_ROOT))

# ---------------------------------------------------- 写入 excel ----------------------------------------------------
def write_batch_excel(results: list, output_root: str) -> str:
    """
    将批量推断结果写入 Excel 文件。

    格式: 每行一个样本, 列 = [类别, 样本名, Precision, Recall, F1, IoU, 预测正类数, 备注]。末尾追加均值汇总行（仅含 GT 的样本）。

    Args:
        - results:     list[dict], run_batch() 的返回值
        - output_root: str,        Excel 写出目录

    Returns:
        - excel_path: str, 写出的 Excel 文件路径
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["类别", "样本名", "Precision", "Recall", "F1", "IoU", "Voxel_Precision", "Voxel_Recall", "Voxel_F1", "Voxel_IoU", "预测正类数", "备注"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    eval_rows = []
    for row in results:
        if row.get("error"):
            ws.append([row["class_folder"], row["sample_name"],
                       "", "", "", "", "", "", "", "", "", f"错误: {row['error']}"])
            continue

        m = row.get("metrics")
        vm = row.get("voxel_metrics")
        if m is not None:
            vp, vr, vf1, viou = ("", "", "", "")
            if vm is not None:
                vp = round(vm.get("precision", 0), 4)
                vr = round(vm.get("recall", 0), 4)
                vf1 = round(vm.get("f1", 0), 4)
                viou = round(vm.get("iou", 0), 4)

            ws.append([
                row["class_folder"], row["sample_name"],
                round(m["precision"], 4), round(m["recall"], 4),
                round(m["f1"], 4), round(m["iou"], 4),
                vp, vr, vf1, viou,
                row.get("num_pred_pos", ""), "",
            ])
            eval_rows.append((m, vm or {}))
        else:
            ws.append([
                row["class_folder"], row["sample_name"],
                "", "", "", "",
                "", "", "", "",
                row.get("num_pred_pos", ""), "无 GT, 仅推断",
            ])

    # 均值汇总行
    if eval_rows:
        ws.append([])
        
        m_list = [r[0] for r in eval_rows]
        vm_list = [r[1] for r in eval_rows if r[1]]
        
        vm_p_mean = round(float(np.mean([x.get("precision", 0) for x in vm_list])), 4) if vm_list else ""
        vm_r_mean = round(float(np.mean([x.get("recall", 0) for x in vm_list])), 4) if vm_list else ""
        vm_f1_mean = round(float(np.mean([x.get("f1", 0) for x in vm_list])), 4) if vm_list else ""
        vm_iou_mean = round(float(np.mean([x.get("iou", 0) for x in vm_list])), 4) if vm_list else ""

        avg_row = [
            "", "【均值】",
            round(float(np.mean([r["precision"] for r in m_list])), 4),
            round(float(np.mean([r["recall"]    for r in m_list])), 4),
            round(float(np.mean([r["f1"]        for r in m_list])), 4),
            round(float(np.mean([r["iou"]       for r in m_list])), 4),
            vm_p_mean, vm_r_mean, vm_f1_mean, vm_iou_mean,
            "", "",
        ]
        ws.append(avg_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    # 自动列宽
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    os.makedirs(output_root, exist_ok=True)
    excel_path = os.path.join(output_root, "results.xlsx")
    wb.save(excel_path)
    print(f"\n[Batch] ✅ Excel 已保存: {excel_path}")
    return excel_path

 
def write_param_search_excel(
    summary: list, param_names: list, output_root: str, output_name: str = None
) -> str:
    """
    将 param_search 汇总结果写入 Excel。

    Args:
        - summary:     list[dict], 按 F1 降序排列的汇总结果
        - param_names: list[str],  被 sweep 的参数名
        - output_root: str,        Excel 写出目录
        - output_name: str | None, Excel 文件名(不含 .xlsx 后缀); None 时使用 "param_search_results"

    Returns:
        - excel_path: str
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ParamSearch"

    headers = param_names + ["avg_Precision", "avg_Recall", "avg_F1", "avg_IoU", "avg_Voxel_Precision", "avg_Voxel_Recall", "avg_Voxel_F1", "avg_Voxel_IoU"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # 颜色标注：第 1 名红色，第 2~5 名黄色
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for i, row in enumerate(summary):
        ws.append(
            [row[k] for k in param_names]
            + [round(row.get("avg_P", 0), 4), round(row.get("avg_R", 0), 4),
               round(row.get("avg_F1", 0), 4), round(row.get("avg_IoU", 0), 4),
               round(row.get("avg_voxel_P", 0), 4), round(row.get("avg_voxel_R", 0), 4),
               round(row.get("avg_voxel_F1", 0), 4), round(row.get("avg_voxel_IoU", 0), 4)]
        )
        if i == 0:
            for cell in ws[ws.max_row]:
                cell.fill = red_fill
        elif 1 <= i <= 4:
            for cell in ws[ws.max_row]:
                cell.fill = yellow_fill

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    # ---- 新增: Per-Sample Sheet ----
    has_per_sample = any("_per_sample" in row for row in summary)
    if has_per_sample:
        ws_ps = wb.create_sheet("Per-Sample")
        headers_ps = param_names + ["sample_name", "Precision", "Recall", "F1", "IoU", "Voxel_Precision", "Voxel_Recall", "Voxel_F1", "Voxel_IoU"]
        ws_ps.append(headers_ps)
        for cell in ws_ps[1]:
            cell.font = Font(bold=True)
            
        for row in summary:
            param_vals = [row[k] for k in param_names]
            ps_records = row.get("_per_sample", [])
            for rec in ps_records:
                ws_ps.append(param_vals + [
                    rec["sample_name"],
                    round(rec["precision"], 4),
                    round(rec["recall"], 4),
                    round(rec["f1"], 4),
                    round(rec["iou"], 4),
                    round(rec.get("voxel_precision", 0), 4),
                    round(rec.get("voxel_recall", 0), 4),
                    round(rec.get("voxel_f1", 0), 4),
                    round(rec.get("voxel_iou", 0), 4)
                ])
                
        for col in ws_ps.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws_ps.column_dimensions[col[0].column_letter].width = max_len + 4

    # ---- 新增: Best-Combination Sheet ----
    if summary and "_per_sample" in summary[0]:
        best_row = summary[0]
        ws_best = wb.create_sheet("Best-Combination")
        
        # 在顶部写入这一组最优的参数组合信息
        ws_best.append(["【最优参数】"] + [f"{k}={best_row[k]}" for k in param_names])
        ws_best.append([])
        
        # 写入表头
        headers_best = ["样本名", "Precision", "Recall", "F1", "IoU", "Voxel_Precision", "Voxel_Recall", "Voxel_F1", "Voxel_IoU"]
        ws_best.append(headers_best)
        for cell in ws_best[ws_best.max_row]:
            cell.font = Font(bold=True)
            
        # 写入最优组合下的逐样本数据
        ps_records = best_row.get("_per_sample", [])
        for rec in ps_records:
            ws_best.append([
                rec.get("sample_name", ""),
                round(rec.get("precision", 0), 4),
                round(rec.get("recall", 0), 4),
                round(rec.get("f1", 0), 4),
                round(rec.get("iou", 0), 4),
                round(rec.get("voxel_precision", 0), 4),
                round(rec.get("voxel_recall", 0), 4),
                round(rec.get("voxel_f1", 0), 4),
                round(rec.get("voxel_iou", 0), 4)
            ])
            
        # 如果有数据，追加一个均值行
        if ps_records:
            ws_best.append([])
            avg_row = [
                "【均值】",
                round(float(np.mean([r["precision"] for r in ps_records])), 4),
                round(float(np.mean([r["recall"]    for r in ps_records])), 4),
                round(float(np.mean([r["f1"]        for r in ps_records])), 4),
                round(float(np.mean([r["iou"]       for r in ps_records])), 4),
                round(float(np.mean([r.get("voxel_precision", 0) for r in ps_records])), 4),
                round(float(np.mean([r.get("voxel_recall", 0)    for r in ps_records])), 4),
                round(float(np.mean([r.get("voxel_f1", 0)        for r in ps_records])), 4),
                round(float(np.mean([r.get("voxel_iou", 0)       for r in ps_records])), 4)
            ]
            ws_best.append(avg_row)
            for cell in ws_best[ws_best.max_row]:
                cell.font = Font(bold=True)
                
        # 自动调整列宽
        for col in ws_best.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=12)
            ws_best.column_dimensions[col[0].column_letter].width = max_len + 4

    os.makedirs(output_root, exist_ok=True)
    # str, Excel 文件名(不含后缀); 未指定则默认 "param_search_results"
    _fname = output_name if output_name else "param_search_results"
    excel_path = os.path.join(output_root, f"{_fname}.xlsx")
    wb.save(excel_path)
    print(f"[ParamSearch] ✅ Excel 已保存: {excel_path}")
    return excel_path


# ---------------------------------------------------- 生成参数网格 ----------------------------------------------------
def generate_param_grid(param_sweep_cfg: list | dict) -> list:
    """
    根据配置生成参数网格（各参数取值的笛卡尔积）。

    Args:
        - param_sweep_cfg: list[dict] 或 dict
            - 若为 list[dict]: 每项含 name / min / max / step，兼容旧接口
            - 若为 dict: 形如 {param_name: {min/max/step}} 或 {param_name: [v1, v2, ...]}

    Returns:
        - param_grid: list[dict], 每项为一组参数组合 {param_name: value}
    """
    all_axes = []
    names = []

    # list[tuple[str, object]], 统一后的参数名与其取值规范
    if isinstance(param_sweep_cfg, dict):
        spec_items = [(str(name), spec) for name, spec in param_sweep_cfg.items()]
    else:
        spec_items = []
        for spec in param_sweep_cfg:
            if not isinstance(spec, dict) or "name" not in spec:
                raise TypeError(
                    "旧版 param_sweep_cfg 的每一项都必须是包含 name 字段的 dict。"
                )
            name = str(spec["name"])
            value_spec = {key: value for key, value in spec.items() if key != "name"}
            spec_items.append((name, value_spec))

    for name, spec in spec_items:
        # list, 当前参数可选值列表
        if isinstance(spec, dict):
            if {"min", "max", "step"} <= set(spec.keys()):
                values = np.arange(spec["min"], spec["max"] + spec["step"] / 2, spec["step"])
                values = np.round(values, 8).tolist()
            elif "values" in spec:
                values = list(spec["values"])
            else:
                raise ValueError(
                    f"[ParamSearch] 参数 {name} 的配置必须包含 min/max/step，或显式提供 values。"
                )
        elif isinstance(spec, (list, tuple)):
            values = list(spec)
        else:
            raise TypeError(
                f"[ParamSearch] 参数 {name} 的配置类型不支持: {type(spec)}"
            )

        all_axes.append(values)
        names.append(name)
        print(f"[ParamSearch] 参数 {name}: {values}")

    combos = list(itertools.product(*all_axes))
    print(f"[ParamSearch] 共 {len(combos)} 组参数组合")
    return [dict(zip(names, combo)) for combo in combos]



# ---------------------------------------------------- 可视化逻辑 ----------------------------------------------------
def _safe_mkdir(dir_path: str) -> None:
    """
    创建目标文件夹，不存在时自动创建。
    """
    if not dir_path:
        return
    os.makedirs(dir_path, exist_ok=True)


def _safe_copy_file(src_path: str, dst_path: str) -> str:
    """
    安全复制文件，自动创建上级路径，复制成功返回目标路径。

    输入参数:
        - src_path: str, 源文件路径
        - dst_path: str, 目标文件路径
    返回:
        - saved_path: str, 实际写出路径；若复制失败则返回 None
    """
    if src_path is None or dst_path is None:
        return None
    if not os.path.exists(src_path):
        return None
    import shutil
    _safe_mkdir(os.path.dirname(dst_path))
    shutil.copy2(src_path, dst_path)
    return dst_path


def _parse_biopython_structure(structure_path: str):
    """
    使用 Biopython 解析 PDB/mmCIF 结构，返回 Structure 对象。

    输入参数:
        - structure_path: str, 输入结构路径（.pdb/.cif/.mmcif）
    返回:
        - structure: Bio.PDB.Structure.Structure, 解析后的结构
    """
    from Bio.PDB import PDBParser, MMCIFParser
    from pathlib import Path as _Path
    if structure_path is None:
        return None
    suffix = _Path(structure_path).suffix.lower()
    if suffix in [".pdb"]:
        parser = PDBParser(QUIET=True)
    elif suffix in [".cif", ".mmcif"]:
        parser = MMCIFParser(QUIET=True)
    else:
        raise ValueError(f"[utils] Unsupported structure format: {suffix}")
    sample_id = _Path(structure_path).stem
    return parser.get_structure(sample_id, structure_path)


def _resolve_model_policy(structure, select_first_model: bool) -> int:
    """
    根据 select_first_model 决定的策略，返回要保留的 model_id（或 None）

    输入参数:
        - structure: Bio.PDB.Structure.Structure, Biopython 结构
        - select_first_model: bool, 是否仅保留第一个 model
    返回:
        - model_id: int | None, 允许写出的 model_id；None 表示全部
    """
    if structure is None:
        return None
    models = list(structure.get_models())
    if not select_first_model and len(models) > 1:
        raise ValueError(
            "[utils] Structure contains multiple models; please set select_first_model=True."
        )
    return 0 if select_first_model else None


def _write_structure_with_select(structure, out_path: str, select_first_model: bool, atom_ids: set):
    """
    将 Biopython 结构以 mmCIF 格式写出，可选写出指定原子

    输入参数:
        - structure: Bio.PDB.Structure.Structure, Biopython 结构
        - out_path: str, 保存路径（.cif）
        - select_first_model: bool, 是否只写第一个 model
        - atom_ids: set[int] | None, 允许写出的原子 id 集合；None 表示全部

    返回:
        - saved_path: str, 实际写出的路径
    """
    if structure is None or out_path is None:
        return None
    from Bio.PDB import MMCIFIO
    from Bio.PDB.PDBIO import Select

    model_id = _resolve_model_policy(structure, select_first_model)

    class _AtomSelect(Select):
        def accept_model(self, model):
            if model_id is None:
                return 1
            return int(model.id) == int(model_id)

        def accept_atom(self, atom):
            if atom_ids is None:
                return 1
            return id(atom) in atom_ids

    _safe_mkdir(os.path.dirname(out_path))
    io = MMCIFIO()
    io.set_structure(structure)
    io.save(out_path, select=_AtomSelect())
    return out_path


def write_structure_as_cif(in_path: str, out_path: str, select_first_model: bool) -> str:
    """
    将 PDB/mmCIF 结构统一写出为 .cif，可选只保留第一个 model

    输入参数:
        - in_path: str, 输入结构路径
        - out_path: str, 输出 .cif 路径
        - select_first_model: bool, 是否仅保留第一个 model
    返回:
        - saved_path: str, 实际写出路径；若输入不存在则返回 None
    """
    if in_path is None or out_path is None:
        return None
    if not os.path.exists(in_path):
        return None
    structure = _parse_biopython_structure(in_path)
    return _write_structure_with_select(structure, out_path, select_first_model, atom_ids=None)


def _collect_receptor_atoms_and_coords(structure, select_first_model: bool):
    """
    对 Biopython 结构执行与 parse_structure 一致的配置过滤，返回原子列表和坐标

    输入:
        - structure: Bio.PDB.Structure.Structure, Biopython 结构
        - select_first_model: bool, 是否只使用第一个 model

    返回:
        - atoms: list[Bio.PDB.Atom.Atom], 按 parse_structure 遍历顺序产生的原子列表
        - coords: np.ndarray, (N, 3), 对应原子坐标（实数）
        - atoms, coords = _collect_receptor_atoms_and_coords(structure, True)
    """
    if structure is None:
        return [], np.empty((0, 3), dtype=np.float32)
    from Bio.PDB.Atom import DisorderedAtom
    from Make_Data.PDB_processor.config import (
        ALLOWED_ELEMENTS,
        Modified_Residues,
        is_protein_residue,
        is_nucleotide_residue,
    )
    from Make_Data.PDB_processor.ligand_candidates import is_connected_to
    from Make_Data.PDB_processor.parser import infer_element_from_atom_name

    model_id = _resolve_model_policy(structure, select_first_model)
    model = structure[0] if model_id is None else structure[int(model_id)]

    atoms = []
    coords = []
    for chain in model:
        for residue in chain:
            het_flag = residue.id[0]
            resname = residue.resname.strip().upper()
            is_modified_receptor = (
                str(het_flag).startswith("H_")
                and resname in Modified_Residues
                and is_connected_to(residue, chain)
            )
            if het_flag != " " and not is_modified_receptor:
                continue
            if not (is_protein_residue(resname) or is_nucleotide_residue(resname)):
                continue

            for atom in residue:
                atom_name = atom.get_name().strip().upper()
                element = None
                if hasattr(atom, "element") and atom.element:
                    element = atom.element.strip().upper()
                    if element == "SE":
                        element = "S"
                    if element not in ALLOWED_ELEMENTS:
                        element = infer_element_from_atom_name(atom_name)
                else:
                    element = infer_element_from_atom_name(atom_name)
                if element is None:
                    continue

                if isinstance(atom, DisorderedAtom):
                    coord = atom.disordered_get_list()[0].get_coord()
                else:
                    coord = atom.get_coord()
                atoms.append(atom)
                coords.append(coord)

    if coords:
        coords = np.array(coords, dtype=np.float32)
    else:
        coords = np.empty((0, 3), dtype=np.float32)
    return atoms, coords


def _map_class_ids(pocket_class_ids: np.ndarray, class_mapping: list):
    """
    对 pocket_class_ids 按 class_mapping 执行重映射，返回映射后结果

    输入:
        - pocket_class_ids: np.ndarray, (N,), 口袋类别 ID
        - class_mapping: list[int] | None, 映射表（old_id -> new_id）

    返回:
        - mapped_ids: np.ndarray, (N,), 映射后 ID
    """
    if class_mapping is None:
        return pocket_class_ids
    mapped = np.zeros_like(pocket_class_ids)
    for old_id, new_id in enumerate(class_mapping):
        mapped[pocket_class_ids == old_id] = new_id
    return mapped


def _collect_ligand_residues(structure, selected_candidates: list, select_first_model: bool):
    """
    根据 LigandCandidate 列表在结构中匹配对应 HETATM 残基

    输入:
        - structure: Bio.PDB.Structure.Structure, Biopython 结构
        - selected_candidates: list[LigandCandidate], 通过筛选的配体候选
        - select_first_model: bool, 是否只使用第一个 model

    返回:
        - residues: set[Bio.PDB.Residue.Residue], 匹配到的残基集合
    """
    if structure is None or not selected_candidates:
        return set()
    model_id = _resolve_model_policy(structure, select_first_model)
    model = structure[0] if model_id is None else structure[int(model_id)]

    residues = set()
    for cand in selected_candidates:
        chain_id = cand.chain_id
        res_id = int(cand.res_id)
        ins_code = str(cand.insertion_code).strip()
        resname = str(cand.resname).strip().upper()
        if chain_id not in model:
            continue
        chain = model[chain_id]
        for residue in chain:
            if residue.resname.strip().upper() != resname:
                continue
            if int(residue.id[1]) != res_id:
                continue
            if str(residue.id[2]).strip() != ins_code:
                continue
            residues.add(residue)
    return residues


def export_selected_ligands_cif(
    cif_gt_path: str,
    out_path: str,
    filter_preset: str,
    select_first_model: bool,
) -> str:
    """
    从真实结构中按 filter_preset 筛选配体，合并写出一个 .cif

    输入:
        - cif_gt_path: str, 真实结构路径（.pdb/.cif/.mmcif）
        - out_path: str, 输出 .cif 路径
        - filter_preset: str, 使用的配体筛选预设（要与工程的 filter_config 一致）
        - select_first_model: bool, 是否只使用第一个 model

    返回:
        - saved_path: str, 实际写出路径；若无配体或文件不存在则返回 None
    """
    if cif_gt_path is None or out_path is None:
        return None
    if not os.path.exists(cif_gt_path):
        return None
    from Make_Data.PDB_processor.parser import parse_structure
    from Make_Data.labels.filter_config import get_filter_preset
    from Make_Data.labels.ligand_filter import filter_and_classify

    parsed = parse_structure(
        cif_gt_path,
        error_dir=str(Path(out_path).parent),
        sample_id=Path(cif_gt_path).stem,
        require_ligand=False,
        select_first_model=select_first_model,
    )
    if parsed is None:
        return None
    filter_config = get_filter_preset(filter_preset)
    if filter_config is None:
        raise ValueError(f"[utils] Unknown filter_preset: {filter_preset}")
    selected, _, _ = filter_and_classify(parsed.ligand_candidates, filter_config)
    if not selected:
        return None

    structure = _parse_biopython_structure(cif_gt_path)
    residues = _collect_ligand_residues(structure, selected, select_first_model)
    if not residues:
        return None

    # 只写配体残基内的原子
    atom_ids = set()
    for res in residues:
        for atom in res.get_atoms():
            atom_ids.add(id(atom))
    return _write_structure_with_select(structure, out_path, select_first_model, atom_ids)


def export_gt_pocket_atoms_cif(
    cif_gt_path: str,
    out_path: str,
    filter_preset: str,
    class_mapping: list,
    select_first_model: bool,
) -> str:
    """
    从真实结构中提取口袋结合原子，并写出 .cif

    输入参数:
        - cif_gt_path: str, 真实结构路径
        - out_path: str, 输出 .cif 路径
        - filter_preset: str, 配体筛选预设
        - class_mapping: list[int] | None, 口袋类别重映射（old_id -> new_id）
        - select_first_model: bool, 是否只使用第一个 model

    返回:
        - saved_path: str, 实际写出路径；若无信息则返回 None

    示例:
        - export_gt_pocket_atoms_cif("gt.cif", "out/gt/pocket_atoms.cif", "five_class", None, True)
    """
    if cif_gt_path is None or out_path is None:
        return None
    if not os.path.exists(cif_gt_path):
        return None
    from Make_Data.PDB_processor.parser import parse_structure
    from Make_Data.labels.filter_config import get_filter_preset
    from Make_Data.labels.ligand_filter import filter_and_classify
    from Make_Data.labels.instance_labels import compute_binding_labels

    parsed = parse_structure(
        cif_gt_path,
        error_dir=str(Path(out_path).parent),
        sample_id=Path(cif_gt_path).stem,
        require_ligand=False,
        select_first_model=select_first_model,
    )
    if parsed is None:
        return None
    filter_config = get_filter_preset(filter_preset)
    if filter_config is None:
        raise ValueError(f"[utils] Unknown filter_preset: {filter_preset}")
    selected, pocket_class_map, _ = filter_and_classify(parsed.ligand_candidates, filter_config)
    if not selected:
        return None

    binding_labels = compute_binding_labels(
        parsed_data=parsed,
        selected_candidates=selected,
        pocket_class_map=pocket_class_map,
        error_dir=None,
        sample_id=Path(cif_gt_path).stem,
        require_binding_site=False,
    )
    if binding_labels is None:
        return None
    pocket_class_ids = binding_labels.get("pocket_class_ids", None)
    if pocket_class_ids is None:
        return None
    pocket_class_ids = _map_class_ids(pocket_class_ids, class_mapping)
    pocket_mask = pocket_class_ids > 0
    if not np.any(pocket_mask):
        return None

    structure = _parse_biopython_structure(cif_gt_path)
    atoms, coords = _collect_receptor_atoms_and_coords(structure, select_first_model)
    if len(atoms) != int(pocket_class_ids.shape[0]):
        return None

    atom_ids = {id(atoms[i]) for i in np.where(pocket_mask)[0].tolist()}
    return _write_structure_with_select(structure, out_path, select_first_model, atom_ids)


def write_point_cloud_cif(
    points: np.ndarray,
    out_path: str,
    atom_name: str = "C",
    element: str = "C",
    resname: str = "POC",
    chain_id: str = "A",
    model_id: int = 1,
) -> str:
    """
    将点云写出为伪原子 mmCIF 文件（不依赖原始结构元数据）。

    输入参数:
        - points: np.ndarray, (N, 3), 点云坐标（世界坐标, Å）
        - out_path: str, 输出 .cif 路径
        - atom_name: str, 伪原子名称前缀（默认 "C"）
        - element: str, 元素符号（默认 "C"）
        - resname: str, 残基名（默认 "POC"）
        - chain_id: str, 链 ID（默认 "A"）
        - model_id: int, 模型编号（默认 1）

    返回:
        - saved_path: str | None, 实际写出路径 out_path；若 points 为空则返回 None
    """
    if out_path is None:
        return None
    if points is None:
        return None

    # np.ndarray, (N, 3), 点云坐标（世界坐标）
    pts = np.asarray(points, dtype=np.float32).reshape(-1, 3)
    if pts.size == 0:
        return None

    _safe_mkdir(os.path.dirname(out_path))

    # list[str], mmCIF 文件内容逐行缓存
    lines = [
        "data_point_cloud",
        "#",
        "loop_",
        "_atom_site.group_PDB",
        "_atom_site.id",
        "_atom_site.type_symbol",
        "_atom_site.label_atom_id",
        "_atom_site.label_alt_id",
        "_atom_site.label_comp_id",
        "_atom_site.label_asym_id",
        "_atom_site.label_entity_id",
        "_atom_site.label_seq_id",
        "_atom_site.Cartn_x",
        "_atom_site.Cartn_y",
        "_atom_site.Cartn_z",
        "_atom_site.occupancy",
        "_atom_site.B_iso_or_equiv",
        "_atom_site.pdbx_PDB_model_num",
    ]

    for i, (x, y, z) in enumerate(pts, start=1):
        atom_id = f"{atom_name}{i}"
        lines.append(
            f"HETATM {i} {element} {atom_id} . {resname} {chain_id} 1 {i} "
            f"{x:.3f} {y:.3f} {z:.3f} 1.00 0.00 {model_id}"
        )

    lines.append("#")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return out_path


def write_grid_as_map(
    data_zyx: np.ndarray,
    out_path: str,
    origin_xyz: np.ndarray,
    voxel_size_xyz: np.ndarray,
) -> str:
    """
    将 (D, H, W) 数组写为 MRC/MAP 文件, 手动设置 voxel_size 和 origin。

    输入参数:
        - data_zyx: np.ndarray, (D, H, W), float32, 体素数据 (ZYX 轴序)
        - out_path: str, 输出文件路径
        - origin_xyz: np.ndarray, (3,), 密度图原点 (x, y, z)
        - voxel_size_xyz: np.ndarray, (3,), 体素大小 (x, y, z)

    输出:
        - saved_path: str, 写出路径; 写出失败返回 None
    """
    if data_zyx is None or out_path is None:
        return None
    import mrcfile
    _safe_mkdir(os.path.dirname(out_path))
    # np.ndarray, (D, H, W), float32, 确保类型正确
    data = np.asarray(data_zyx, dtype=np.float32)
    # np.ndarray, (3,), float64
    origin = np.asarray(origin_xyz, dtype=np.float64).ravel()
    voxel = np.asarray(voxel_size_xyz, dtype=np.float64).ravel()
    try:
        with mrcfile.new(out_path, overwrite=True) as mrc:
            mrc.set_data(data)
            # 设置 origin (x, y, z)
            mrc.header.origin.x = float(origin[0])
            mrc.header.origin.y = float(origin[1])
            mrc.header.origin.z = float(origin[2])
            # 设置 nxstart/nystart/nzstart 为 0
            mrc.header.nxstart = 0
            mrc.header.nystart = 0
            mrc.header.nzstart = 0
            # 设置 cella (单位格子的绝对尺寸 = 网格维度 × voxel_size)
            # data shape 顺序为 (Z, Y, X), cella 顺序为 (X, Y, Z)
            nz, ny, nx = data.shape
            mrc.header.cella.x = float(nx * voxel[0])
            mrc.header.cella.y = float(ny * voxel[1])
            mrc.header.cella.z = float(nz * voxel[2])
            mrc.update_header_stats()
        return out_path
    except Exception as e:
        print(f"[write_grid_as_map] 写出失败: {out_path}, {e}")
        return None


def write_pymol_vis_script(
    vis_paths: dict,
    instance_label: np.ndarray,
    out_path: str,
) -> str:
    """
    生成 PyMOL 启动脚本, 打开后自动加载当前样本可视化文件并按预测 instance 建组。

    输入参数:
        - vis_paths: dict, build_infer_vis_bundle() 已写出的路径字典, 包含 root_dir/gt_structure/pred_instances 等路径
        - instance_label: np.ndarray | None, (D, H, W), 预测 instance 标签, 0 为背景, 正整数为 instance_id
        - out_path: str, .pml 脚本输出路径

    使用说明:
        - 在 PyMOL 中打开 out_path, 不要单独打开 pred/instances_pred.cif。
        - 脚本会把每个 instance_id 创建为 pred_instance_### 对象, 并加入 pred_instances group。
        - 若同时存在 gt/pred 结构、配体、口袋原子和密度图, 脚本会一并加载并放入 gt/pred/pred_density group。

    输出:
        - saved_path: str, 写出的 PyMOL 脚本路径
    """
    script_dir = os.path.dirname(out_path)
    _safe_mkdir(script_dir)

    def _rel_pml_path(file_path: str) -> str:
        rel_path = os.path.relpath(file_path, script_dir)
        return rel_path.replace("\\", "/").replace('"', '\\"')

    def _append_load(
        lines: list[str],
        path_key: str,
        object_name: str,
        group_name: str,
        show_commands: list[str],
    ) -> None:
        file_path = vis_paths.get(path_key)
        if file_path is None or not os.path.exists(file_path):
            return
        lines.append(f'load "{_rel_pml_path(file_path)}", {object_name}')
        lines.extend(show_commands)
        lines.append(f"group {group_name}, {object_name}")

    # list[str], PyMOL 脚本逐行命令
    lines = [
        "reinitialize",
        "set auto_zoom, off",
        "bg_color white",
        "set retain_order, 1",
    ]
    _append_load(lines, "gt_structure", "gt_structure", "gt", ["hide everything, gt_structure", "show cartoon, gt_structure", "color gray70, gt_structure"])
    _append_load(lines, "gt_ligand", "gt_ligand", "gt", ["show sticks, gt_ligand", "color yellow, gt_ligand"])
    _append_load(lines, "gt_pocket_atoms", "gt_pocket_atoms", "gt", ["show spheres, gt_pocket_atoms", "set sphere_scale, 0.35, gt_pocket_atoms", "color cyan, gt_pocket_atoms"])
    _append_load(lines, "gt_density", "gt_density_raw", "gt_density", [])
    _append_load(lines, "gt_density_resampled", "gt_density_resampled_map", "gt_density", [])
    _append_load(lines, "pred_structure", "pred_structure", "pred", ["hide everything, pred_structure", "show cartoon, pred_structure", "color gray60, pred_structure"])
    _append_load(lines, "pred_pocket_atoms", "pred_pocket_atoms", "pred", ["show spheres, pred_pocket_atoms", "set sphere_scale, 0.35, pred_pocket_atoms", "color orange, pred_pocket_atoms"])
    _append_load(lines, "pred_density_binary", "pred_density_binary_map", "pred_density", [])
    _append_load(lines, "pred_density", "pred_density_masked_map", "pred_density", [])
    _append_load(lines, "pred_density_prob", "pred_density_prob_map", "pred_density", [])

    # list[int], 预测 instance_id 列表, 仅保留正整数标签
    instance_ids = []
    if instance_label is not None:
        label = np.asarray(instance_label, dtype=np.int32)
        instance_ids = [int(v) for v in np.unique(label) if int(v) > 0]

    instances_path = vis_paths.get("pred_instances")
    if instances_path is not None and os.path.exists(instances_path):
        if instance_ids:
            colors = ["tv_red", "tv_green", "tv_blue", "yellow", "cyan", "magenta", "orange", "lime", "marine", "salmon", "purple", "wheat"]
            lines.append(f'load "{_rel_pml_path(instances_path)}", pred_instances_raw')
            lines.append("hide everything, pred_instances_raw")
            for i, instance_id in enumerate(instance_ids):
                instance_obj = f"pred_instance_{instance_id:03d}"
                instance_color = colors[i % len(colors)]
                lines.append(f"create {instance_obj}, pred_instances_raw and resi {instance_id}")
                lines.append(f"show spheres, {instance_obj}")
                lines.append(f"set sphere_scale, 0.35, {instance_obj}")
                lines.append(f"color {instance_color}, {instance_obj}")
                lines.append(f"group pred_instances, {instance_obj}")
            lines.append("delete pred_instances_raw")
            lines.append("group pred, pred_instances")
            lines.append("zoom pred_instances")
        else:
            lines.append(f'load "{_rel_pml_path(instances_path)}", pred_instances')
            lines.append("show spheres, pred_instances")
            lines.append("set sphere_scale, 0.35, pred_instances")
            lines.append("color orange, pred_instances")
            lines.append("group pred, pred_instances")

    lines.append("orient")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    return out_path


def build_infer_vis_bundle(
    output_root: str,
    cif_path: str,
    map_path: str,
    cif_gt_path: str,
    pred_atom_coords: np.ndarray,
    prob_threshold: float,  # 可为空, 仅用于写文件名写标签
    filter_preset: str,
    class_mapping: list,
    pdb_id: str,
    select_first_model: bool,
    pred_voxel_mask: np.ndarray = None,
    resampled_emdb: np.ndarray = None,
    origin: np.ndarray = None,
    voxel_size: np.ndarray = None,
    pred_voxel_prob: np.ndarray = None,
    pred_instance_label: np.ndarray = None,
    write_pred_atom_coords: bool = True,
) -> dict:
    """
    根据推断输入/输出打包可视化文件夹，结构为 output_root/<pdb_id>/{gt,pred}

    gt/ 文件名加 _gt 后缀, pred/ 文件名加 _pred 后缀, 避免 PyMOL 同名冲突。
    若需要在 PyMOL 中直接看到预测 instance 分组, 打开 output_root/<pdb_id>/open_in_pymol.pml。

    输入参数:
        - output_root: str, 输出根目录
        - cif_path: str, 推断输入结构路径（预测结构可为 AF3/CryoAtom 输出）
        - map_path: str | None, 密度图路径
        - cif_gt_path: str | None, 真实结构路径（可为 None，表示该样本没有独立 GT）
        - pred_atom_coords: np.ndarray, (N_pred, 3), 预测为正类的原子点云（世界坐标）
        - prob_threshold: float, 预测口袋概率阈值
        - filter_preset: str, 配体筛选预设
        - class_mapping: list[int] | None, 口袋类别重映射
        - pdb_id: str | None, 样本 ID（若 None 则从 cif_gt_path/cif_path 推断）
        - select_first_model: bool, 是否只使用第一个 model
        - pred_voxel_mask: np.ndarray | None, (D, H, W), int64, 预测正类体素 mask (重采样空间)
        - resampled_emdb: np.ndarray | None, (D, H, W), float32, 重采样后 EMDB 密度 (第一通道)
        - origin: np.ndarray | None, (3,), 重采样后密度图原点 (x, y, z)
        - voxel_size: np.ndarray | None, (3,), 重采样后体素大小 (x, y, z)
        - pred_voxel_prob: np.ndarray | None, (D, H, W), float32, ligand 概率图
        - pred_instance_label: np.ndarray | None, (D, H, W), int32, 预测 instance 标签
        - write_pred_atom_coords: bool, 是否写出 pred_atom_coords 对应的预测原子点云

    返回:
        - result: dict, 各个输出文件路径的汇总字典
    
    调用模式:
        - voxel_single: run_voxel_single() 在 vis_enable=True 且 vis_output_root 非空时调用本函数。
        - voxel_batch: run_voxel_batch() 会逐样本转成 voxel_batch_item, 每个样本沿用 voxel_single 的可视化输出。
        - voxel_param_search: 构建缓存和逐组调参不调用本函数; 仅在 best_outputs 阶段按最优后处理参数可选调用。

    保存内容:
        - GT 类:
            - gt/structure_gt.cif: 由 cif_gt_path 写出; 若未提供 cif_gt_path, 则由 cif_path 写出。
            - gt/density_gt.map: 复制原始 map_path 对应密度图; map_path 为空时不写出。
            - gt/density_resampled_gt.map: 写出重采样后的 EMDB 密度; 需要 resampled_emdb、origin、voxel_size 同时存在。
            - gt/ligand_gt.cif: 从 GT 结构中按 filter_preset 筛选配体后写出; 无配体或无结构时不写出。
            - gt/pocket_atoms_gt.cif: 从 GT 结构中写出真实口袋原子; 无 GT 口袋标签时不写出。

        - PRED 类:
            - pred/structure_pred.cif: 同时提供 cif_path 和 cif_gt_path 时, 将 cif_path 作为预测结构写出。
            - pred/pocket_atoms_th*_pred.cif: write_pred_atom_coords=True 时写出 pred_atom_coords 伪原子点云。
            - pred/density_pred_binary.map: 写出后处理后的预测正类体素二值 mask; voxel_pipeline 的 single/batch/cache 可视化会写出。
            - pred/density_pred.map: 写出预测正类区域的连续密度(mask × resampled_emdb); 需要 resampled_emdb 同时存在。
            - pred/density_pred_prob.map: 写出模型 ligand 概率图; voxel_pipeline 的 single/batch/cache 可视化会写出。
            - pred/instances_pred.cif: 将预测 instance_label 的体素中心写成伪原子 CIF; voxel_pipeline 的 single/batch/cache 可视化会写出。
            - open_in_pymol.pml: PyMOL 启动脚本, 自动加载可视化文件并把预测 instance 拆成 pred_instance_### 分组对象。
    """
    result = {}

    # -------- 0. 样本 ID --------
    if not pdb_id:
        if cif_gt_path:
            pdb_id = Path(cif_gt_path).stem
        else:
            pdb_id = Path(cif_path).stem

    root_dir = os.path.join(output_root, pdb_id)
    gt_dir = os.path.join(root_dir, "gt")
    pred_dir = os.path.join(root_dir, "pred")
    _safe_mkdir(gt_dir)
    _safe_mkdir(pred_dir)
    result["root_dir"] = root_dir
    result["gt_dir"] = gt_dir
    result["pred_dir"] = pred_dir

    # -------- 1. GT: 结构 + 密度图 --------
    gt_source = cif_gt_path if cif_gt_path else cif_path
    gt_struct_path = os.path.join(gt_dir, "structure_gt.cif")
    result["gt_structure"] = write_structure_as_cif(gt_source, gt_struct_path, select_first_model)

    if map_path:
        suffix = "".join(Path(map_path).suffixes)
        density_name = f"density_gt{suffix if suffix else '.map'}"
        density_out = os.path.join(gt_dir, density_name)
        result["gt_density"] = _safe_copy_file(map_path, density_out)
    else:
        result["gt_density"] = None

    # -------- 1b. GT: 重采样后密度图 --------
    if resampled_emdb is not None and origin is not None and voxel_size is not None:
        resampled_density_out = os.path.join(gt_dir, "density_resampled_gt.map")
        result["gt_density_resampled"] = write_grid_as_map(
            resampled_emdb, resampled_density_out, origin, voxel_size,
        )
    else:
        result["gt_density_resampled"] = None

    # -------- 2. GT: 纯配体 + 口袋原子 --------
    gt_ligand_out = os.path.join(gt_dir, "ligand_gt.cif")
    result["gt_ligand"] = export_selected_ligands_cif(
        gt_source,
        gt_ligand_out,
        filter_preset,
        select_first_model,
    )
    gt_pocket_out = os.path.join(gt_dir, "pocket_atoms_gt.cif")
    result["gt_pocket_atoms"] = export_gt_pocket_atoms_cif(
        gt_source,
        gt_pocket_out,
        filter_preset,
        class_mapping,
        select_first_model,
    )

    # -------- 3. Pred: 模拟结构（如有） --------
    pred_struct_path = os.path.join(pred_dir, "structure_pred.cif")
    if cif_path and cif_gt_path and os.path.exists(cif_path):  # 只有明确存在(预测结构, 真实结构)时, 才把cif_path当作预测结构并写入预测结构文件夹
        result["pred_structure"] = write_structure_as_cif(cif_path, pred_struct_path, select_first_model)
    else:
        result["pred_structure"] = None

    # -------- 4. Pred: 预测原子点云 --------
    if write_pred_atom_coords:
        th_tag = f"{prob_threshold:.3f}" if prob_threshold is not None else "na"
        pred_pocket_out = os.path.join(pred_dir, f"pocket_atoms_th{th_tag}_pred.cif")
        result["pred_pocket_atoms"] = write_point_cloud_cif(
            pred_atom_coords,
            pred_pocket_out,
        )
    else:
        result["pred_pocket_atoms"] = None

    # -------- 5. Pred: 预测正类体素密度图 --------
    if pred_voxel_mask is not None and origin is not None and voxel_size is not None:
        # 5a. 始终生成二值 mask 可视化 (0/1)
        pred_binary_out = os.path.join(pred_dir, "density_pred_binary.map")
        result["pred_density_binary"] = write_grid_as_map(
            pred_voxel_mask.astype(np.float32),
            pred_binary_out,
            origin,
            voxel_size,
        )

        # 5b. 有 resampled_emdb 时, 额外生成连续密度可视化 (mask × EMDB)
        if resampled_emdb is not None:
            pred_density_out = os.path.join(pred_dir, "density_pred.map")
            # np.ndarray, (D, H, W), float32, 预测区域的连续密度值
            pred_density_data = pred_voxel_mask.astype(np.float32) * resampled_emdb.astype(np.float32)
            result["pred_density"] = write_grid_as_map(
                pred_density_data,
                pred_density_out,
                origin,
                voxel_size,
            )
        else:
            result["pred_density"] = None
    else:
        result["pred_density_binary"] = None
        result["pred_density"] = None

    # -------- 6. Pred: voxel-only 概率图与 instance 点云 --------
    if pred_voxel_prob is not None and origin is not None and voxel_size is not None:
        pred_prob_out = os.path.join(pred_dir, "density_pred_prob.map")
        result["pred_density_prob"] = write_grid_as_map(
            np.asarray(pred_voxel_prob, dtype=np.float32),
            pred_prob_out,
            origin,
            voxel_size,
        )
    else:
        result["pred_density_prob"] = None

    if pred_instance_label is not None and origin is not None and voxel_size is not None:
        instances_out = os.path.join(pred_dir, "instances_pred.cif")
        # np.ndarray, (D, H, W), 预测 instance 标签, 0 为背景
        pred_instance_label_array = np.asarray(pred_instance_label, dtype=np.int32)
        result["pred_instances"] = write_instance_voxel_centers_cif(
            instance_label=pred_instance_label_array,
            origin=origin,
            voxel_size=voxel_size,
            out_path=instances_out,
        )
        result["pymol_script"] = write_pymol_vis_script(
            vis_paths=result,
            instance_label=pred_instance_label_array,
            out_path=os.path.join(root_dir, "open_in_pymol.pml"),
        )
    else:
        result["pred_instances"] = None
        result["pymol_script"] = write_pymol_vis_script(
            vis_paths=result,
            instance_label=None,
            out_path=os.path.join(root_dir, "open_in_pymol.pml"),
        )

    return result


def write_instance_voxel_centers_cif(
    instance_label: np.ndarray,
    origin: np.ndarray,
    voxel_size: np.ndarray,
    out_path: str,
) -> str:
    """
    将预测 instance 的体素中心写成 pseudo atom CIF。

    输入参数:
        - instance_label: np.ndarray, (D,H,W), int32, 预测 instance 标签, 0为背景
        - origin: np.ndarray, (3,), 世界坐标原点(x,y,z)
        - voxel_size: np.ndarray, (3,), 体素大小(x,y,z)
        - out_path: str, CIF 输出路径

    输出:
        - out_path: str, 写出的 CIF 文件路径
    """
    _safe_mkdir(os.path.dirname(out_path))
    label = np.asarray(instance_label, dtype=np.int32)
    origin = np.asarray(origin, dtype=np.float32).reshape(3)
    voxel_size = np.asarray(voxel_size, dtype=np.float32).reshape(3)

    lines = [
        "data_pred_instances",
        "#",
        "loop_",
        "_atom_site.group_PDB",
        "_atom_site.id",
        "_atom_site.type_symbol",
        "_atom_site.label_atom_id",
        "_atom_site.label_comp_id",
        "_atom_site.label_asym_id",
        "_atom_site.label_seq_id",
        "_atom_site.Cartn_x",
        "_atom_site.Cartn_y",
        "_atom_site.Cartn_z",
        "_atom_site.occupancy",
        "_atom_site.B_iso_or_equiv",
    ]
    atom_id = 1
    for instance_id in [int(v) for v in np.unique(label) if int(v) > 0]:
        coords_zyx = np.argwhere(label == instance_id)
        for coord_zyx in coords_zyx:
            z, y, x = [int(v) for v in coord_zyx.tolist()]
            world_x = float(origin[0] + (x + 0.5) * voxel_size[0])
            world_y = float(origin[1] + (y + 0.5) * voxel_size[1])
            world_z = float(origin[2] + (z + 0.5) * voxel_size[2])
            lines.append(
                f"HETATM {atom_id} C C LIG A {instance_id} "
                f"{world_x:.3f} {world_y:.3f} {world_z:.3f} 1.00 0.00"
            )
            atom_id += 1
    lines.append("#")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return out_path


def write_voxel_batch_excel(results: list[dict], output_root: str) -> str:
    """
    将 voxel-only batch 推理结果写入 Excel。

    输入参数:
        - results: list[dict], 可变长度, 每个样本的推理/评估结果
        - output_root: str, 输出目录

    输出:
        - excel_path: str, 写出的 Excel 文件路径
    """
    os.makedirs(output_root, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VoxelResults"
    headers = [
        "sample_name",
        "num_candidates",
        "voxel_precision",
        "voxel_recall",
        "voxel_f1",
        "voxel_iou",
        "voxel_dice",
        "instance_precision",
        "instance_recall",
        "instance_f1",
        "num_pred_instances",
        "num_gt_instances",
        "error",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    metric_rows: list[dict] = []
    for row in results:
        if row.get("error"):
            ws.append([row.get("sample_name", ""), "", "", "", "", "", "", "", "", "", "", "", row["error"]])
            continue
        metrics = row.get("metrics", {}) or {}
        ws.append(
            [
                row.get("sample_name", ""),
                row.get("num_candidates", ""),
                metrics.get("voxel_precision", ""),
                metrics.get("voxel_recall", ""),
                metrics.get("voxel_f1", ""),
                metrics.get("voxel_iou", ""),
                metrics.get("voxel_dice", ""),
                metrics.get("instance_precision", ""),
                metrics.get("instance_recall", ""),
                metrics.get("instance_f1", ""),
                metrics.get("num_pred_instances", ""),
                metrics.get("num_gt_instances", ""),
                "",
            ]
        )
        if metrics:
            metric_rows.append(metrics)

    if metric_rows:
        ws.append([])
        summary = ["MEAN", ""]
        for metric_name in headers[2:12]:
            values = [float(item[metric_name]) for item in metric_rows if metric_name in item]
            summary.append(float(np.mean(values)) if values else "")
        summary.append("")
        ws.append(summary)

    excel_path = os.path.join(output_root, "voxel_batch_results.xlsx")
    wb.save(excel_path)
    return excel_path
